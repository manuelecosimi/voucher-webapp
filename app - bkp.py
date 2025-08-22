from flask import Flask, render_template, request
import pandas as pd
from datetime import datetime
import os

from openpyxl import load_workbook

app = Flask(__name__)

# --- Path robusto all'Excel ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, 'voucher-clienti.xlsx')

# --- Utility ---

def _parse_money(val):
    """Converte stringhe tipo '€ 1.234,56' in float 1234.56. Ritorna None se vuoto/non valido."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return None
    txt = str(val)
    txt = txt.replace('€', '').replace('\xa0', '').replace('\u202f', '').strip()
    if ',' in txt and '.' in txt:
        txt = txt.replace('.', '')
    txt = txt.replace(',', '.')
    try:
        return float(txt) if txt else None
    except Exception:
        return None

def format_valore(valore):
    num = _parse_money(valore)
    if num is None:
        return ""
    return f"€{int(num)}" if float(num).is_integer() else f"€{num:.2f}"

# --- Ricerca voucher ---

def cerca_voucher(numero):
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    headers = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    ordine_col = headers.index("ORDINE") + 1

    found_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        ordine_val = row[ordine_col - 1].value
        if str(ordine_val).strip() == numero:
            found_row = row
            break

    if not found_row:
        wb.close()
        return None

    values = dict(zip(headers, [cell.value for cell in found_row]))
    wb.close()

    # Storico/scalature
    storico = []
    somma_scalature = 0.0
    for col_name in ['1', '2', '3', '4', '5']:
        val = values.get(col_name)
        imp = _parse_money(val)
        if imp is not None:
            storico.append(format_valore(imp))
            somma_scalature += imp

    # Valore totale e residuo
    valore_raw = _parse_money(values.get('VALORE')) or 0.0
    residuo_raw = max(0.0, valore_raw - somma_scalature)

    # Non utilizzabile se tutte le 5 scalature sono valorizzate
    def _filled(x):
        if x is None:
            return False
        if isinstance(x, str) and x.strip() == "":
            return False
        num = _parse_money(x)
        return num is not None and num != 0
    non_utilizzabile = all(_filled(values.get(c)) for c in ['1', '2', '3', '4', '5'])

    # --- NEW: mappa note per appuntamento (in NOTE salviamo "[N] testo") ---
    notes_map = {}
    note_raw = values.get('NOTE') or ""
    if isinstance(note_raw, str):
        import re
        for m in re.finditer(r"\[(\d)\]\s*(.+?)(?=(?:\s*\[\d\])|$)", note_raw, flags=re.S):
            idx = int(m.group(1))
            txt = m.group(2).strip().replace("\r", " ").replace("\n", " ")
            notes_map[idx] = txt

    return {
        'numero': numero,
        'ordine': values.get('ORDINE'),
        'status': "scaduta" if residuo_raw == 0 else "attiva",
        'valore': format_valore(valore_raw),
        'residuo': format_valore(residuo_raw),
        'servizio': values.get('SERVIZIO') or "",
        'card_fisica': "✅" if values.get('N° CARD') else "",
        'box': "✅" if values.get('BOX') else "",
        'card': values.get('CARD') or "",
        'email': values.get('CLIENTE \\ MAIL ORDINE'),
        'data': values.get('DATA').strftime("%d/%m/%Y") if values.get('DATA') else "",
        'storico': storico,
        'note': values.get('NOTE') or "",
        'storico_note_map': notes_map,   # --- NEW ---
        'non_utilizzabile': non_utilizzabile
    }

# --- ROUTES ---

@app.route('/', methods=['GET', 'POST'])
def index():
    errore = None
    if request.method == 'POST':
        numero = request.form.get('numero', '').strip()
        if numero:
            if not numero.startswith('#'):
                numero = f"#{numero}"
            risultato = cerca_voucher(numero)
            if risultato:
                return render_template('voucher.html', voucher=risultato)
            else:
                errore = "Voucher non trovato. Controlla il numero inserito."
        else:
            errore = "Inserisci un numero di voucher valido."
    return render_template('index.html', errore=errore)

@app.route('/gestisci', methods=['GET', 'POST'])
def gestisci():
    numero = request.args.get('numero', '').strip()
    if not numero:
        return "Numero voucher mancante"

    if numero.startswith('##'):
        numero = numero[1:]

    df = pd.read_excel(EXCEL_PATH)
    df.columns = df.columns.str.strip()

    riga = df[df['ORDINE'].astype(str).str.strip() == numero]
    if riga.empty:
        return "Voucher non trovato"

    index = riga.index[0]
    r = riga.iloc[0]

    # Prima colonna scalatura libera
    prossimo = None
    for i, col in enumerate(['1', '2', '3', '4', '5'], start=1):
        if pd.isna(r[col]) or (isinstance(r[col], str) and r[col].strip() == ""):
            prossimo = i
            break

    non_utilizzabile = prossimo is None
    label_appuntamento = f"Appuntamento {prossimo}" if prossimo else ""

    if request.method == 'POST':
        if non_utilizzabile:
            return "Voucher non più utilizzabile (scalature esaurite)"

        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        excel_row = index + 2  # +1 header, +1 base 1

        # Aggiorna N° Card
        ws[f'B{excel_row}'] = request.form.get('card', '')

        nota_form = (request.form.get('note') or '').strip()

        if not pd.isna(r['SERVIZIO']) and str(r['SERVIZIO']).strip() != "":
            # Caso "Servizio effettuato" (checkbox) -> copia VALORE nella prima scalatura libera
            if request.form.get('servizio_effettuato'):
                valore = _parse_money(r['VALORE']) or 0.0
                col_letter = ['J', 'K', 'L', 'M', 'N'][prossimo - 1]
                ws[f'{col_letter}{excel_row}'] = valore
            # Nota generica: appendi (senza tag appuntamento)
            if nota_form:
                existing = ws[f'P{excel_row}'].value or ''
                sep = '\n' if existing else ''
                ws[f'P{excel_row}'] = f"{existing}{sep}{nota_form}"
        else:
            # Caso "scalatura manuale" (Appuntamento N)
            importo_txt = request.form.get('scalatura', '').strip()
            imp = _parse_money(importo_txt)
            if imp is None or imp <= 0:
                wb.close()
                return "Importo non valido"
            col_letter = ['J', 'K', 'L', 'M', 'N'][prossimo - 1]
            ws[f'{col_letter}{excel_row}'] = imp

            # --- NEW: Appendi nota taggata con numero appuntamento ---
            if nota_form:
                existing = ws[f'P{excel_row}'].value or ''
                sep = '\n' if existing else ''
                ws[f'P{excel_row}'] = f"{existing}{sep}[{prossimo}] {nota_form}"

        wb.save(EXCEL_PATH)
        wb.close()
        return render_template('voucher.html', voucher=cerca_voucher(numero))

    # GET: prepara dati per il template gestisci
    return render_template(
        'gestisci.html',
        label_appuntamento=label_appuntamento,
        numero=numero,
        voucher={
            'numero': numero,
            'card': (r['N° CARD'] if not pd.isna(r['N° CARD']) else ''),
            'note': (r['NOTE'] if not pd.isna(r['NOTE']) else ''),
            'servizio': (r['SERVIZIO'] if not pd.isna(r['SERVIZIO']) else ''),
            'valore': format_valore(r['VALORE']),
            'non_utilizzabile': non_utilizzabile,
            'prossimo_appuntamento': prossimo,
            'colonna_attiva': prossimo
        }
    )

if __name__ == '__main__':
    app.run(debug=True)
