from flask import Flask, render_template, request
import pandas as pd
from datetime import datetime, timedelta
import os, time
import re

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# OneDrive / MS Graph (usato solo se USE_CLOUD=1)
import requests
import msal
from dotenv import load_dotenv

# -------------------- ENV & MODALITÀ DEV --------------------
env_file = ".env.dev" if os.path.exists(".env.dev") else ".env"
load_dotenv(env_file)

DEV_MODE  = os.getenv("DEV_MODE")  == "1"
USE_CLOUD = os.getenv("USE_CLOUD") == "1"
SKIP_CLOUD = os.getenv("SKIP_CLOUD") == "1"
# ------------------------------------------------------------

app = Flask(__name__)

# ---------- HEALTH CHECK PER RENDER ----------
@app.get("/healthz")
def healthz():
    return "ok", 200
# --------------------------------------------

# --- Path robusto all'Excel ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

if DEV_MODE and not USE_CLOUD:
    EXCEL_PATH = os.path.join(BASE_DIR, os.getenv("LOCAL_EXCEL_PATH", "voucher-clienti.xlsx"))
else:
    EXCEL_PATH = os.path.join(BASE_DIR, "voucher-clienti.xlsx")

# >>> Client OneDrive usato solo se USE_CLOUD=1
graph = None
ONEDRIVE_PATH = None
if USE_CLOUD:
    from graph_client import GraphClient
    ONEDRIVE_PATH = os.getenv("ONEDRIVE_EXCEL_PATH", "/voucher-clienti.xlsx")
    graph = GraphClient()

def sync_from_cloud():
    if SKIP_CLOUD or not USE_CLOUD or graph is None:
        return
    try:
        graph.download_excel(EXCEL_PATH, ONEDRIVE_PATH)
    except Exception as e:
        print(f"[SYNC] download da OneDrive saltato: {e}")

def sync_to_cloud() -> bool:
    if SKIP_CLOUD or not USE_CLOUD or graph is None:
        return True
    try:
        graph.upload_excel(EXCEL_PATH, ONEDRIVE_PATH)
        print("[SYNC] upload completato su OneDrive")
        return True
    except Exception as e:
        print(f"[SYNC] upload verso OneDrive fallito/ritardato: {e}")
        return False

# --- Utility ---

def _parse_money(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return None
    txt = str(val)
    txt = txt.replace('€', '').replace('\xa0', '').replace('\u202f', '').strip()
    if ',' in txt and '.' in txt:
        txt = txt.replace('.', '')
    txt = txt.replace(',', '.')
    try:
        return float(txt) if txt else None
    except Exception:
        return None

def format_valore(valore):
    num = _parse_money(valore)
    if num is None:
        return ""
    return f"€{int(num)}" if float(num).is_integer() else f"€{num:.2f}"

def _norm_card(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

def _digits(s) -> str:
    return "".join(re.findall(r"\d+", str(s))) if s is not None else ""

def find_service_col(columns):
    for c in columns:
        if c is None:
            continue
        if str(c).strip().upper().startswith("SERVIZIO"):
            return c
    return None

def find_note_col(columns):
    for c in columns:
        if c is None:
            continue
        s = str(c).strip().upper()
        if "NOTE" in s:
            return c
    return None

def find_first_empty_row(ws, start_row=2, key_col=1):
    row = start_row
    while True:
        val = ws.cell(row=row, column=key_col).value
        if val is None or str(val).strip() == "":
            return row
        row += 1


def get_next_manual_gift_order_number(ws):
    max_num = 0

    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        raw = row[0]
        if raw is None:
            continue

        s = str(raw).strip()
        digits = _digits(s)

        if digits.startswith("9"):
            try:
                n = int(digits)
                if n > max_num:
                    max_num = n
            except Exception:
                pass

    if max_num == 0:
        max_num = 90000

    return f"#{max_num + 1}"


def gift_number_exists(ws, gift_number):
    target = str(gift_number).strip().zfill(4)

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        raw = row[0]
        if raw is None:
            continue

        s = str(raw).strip()
        if s.endswith(".0"):
            s = s[:-2]

        if s.isdigit():
            s = s.zfill(4)

        if s == target:
            return True

    return False


def copy_row_style(ws, source_row, target_row, max_col=None):
    if max_col is None:
        max_col = ws.max_column

    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = copy(source_cell.number_format)

        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)

        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


# --- Ricerca voucher ---
def cerca_voucher(numero, force_local: bool = False):
    if not force_local:
        sync_from_cloud()

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    headers = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    ordine_col = headers.index("ORDINE") + 1

    servizio_col = find_service_col(headers)
    note_col = find_note_col(headers)

    target = _digits(numero)

    found_row = None
    for row in ws.iter_rows(min_row=2, values_only=False):
        ordine_val = row[ordine_col - 1].value
        if _digits(ordine_val) == target:
            found_row = row
            break

    if not found_row:
        wb.close()
        return None

    values = dict(zip(headers, [cell.value for cell in found_row]))
    wb.close()

    # Storico/scalature
    storico = []
    somma_scalature = 0.0
    for col_name in ['1', '2', '3', '4', '5']:
        val = values.get(col_name)
        imp = _parse_money(val)
        if imp is not None:
            storico.append(format_valore(imp))
            somma_scalature += imp

    # Valore totale e residuo
    valore_raw = _parse_money(values.get('VALORE')) or 0.0
    residuo_raw = max(0.0, valore_raw - somma_scalature)

    # Non utilizzabile se tutte le 5 scalature sono valorizzate
    def _filled(x):
        if x is None:
            return False
        if isinstance(x, str) and x.strip() == "":
            return False
        num = _parse_money(x)
        return num is not None and num != 0
    non_utilizzabile = all(_filled(values.get(c)) for c in ['1', '2', '3', '4', '5'])

    # servizio sicuro
    servizio_val = values.get(servizio_col) if servizio_col else ""

    # NOTE e DATA sicure
    note_raw = values.get(note_col) or ""
    _data_cell = values.get('DATA')

    _data_str = ""
    try:
        if hasattr(_data_cell, "strftime"):
            _data_str = _data_cell.strftime("%d/%m/%Y")
        elif isinstance(_data_cell, (int, float)) and _data_cell:
            excel_epoch = datetime(1899, 12, 30)
            _data_str = (excel_epoch + timedelta(days=float(_data_cell))).strftime("%d/%m/%Y")
        elif isinstance(_data_cell, str):
            _data_str = _data_cell.strip()
    except Exception:
        _data_str = str(_data_cell).strip() if _data_cell else ""

    notes_map = {}
    if isinstance(note_raw, str):
        for m in re.finditer(r"\[(\d)\]\s*(.+?)(?=(?:\s*\[\d\])|$)", note_raw, flags=re.S):
            idx = int(m.group(1))
            txt = m.group(2).strip().replace("\r", " ").replace("\n", " ")
            notes_map[idx] = txt

    return {
        'numero': target,
        'ordine': values.get('ORDINE'),
        'status': "scaduta" if residuo_raw == 0 else "attiva",
        'valore': format_valore(valore_raw),
        'residuo': format_valore(residuo_raw),
        'servizio': servizio_val or "",
        'card_fisica': "✅" if values.get('N° CARD') else "",
        'box': "✅" if values.get('BOX') else "",
        'card': values.get('CARD') or "",
        'email': values.get('CLIENTE \\ MAIL ORDINE'),
        'data': _data_str,
        'storico': storico,
        'note': note_raw or "",
        'storico_note_map': notes_map,
        'non_utilizzabile': non_utilizzabile
    }

# --- ROUTES ---

@app.route('/', methods=['GET', 'POST'])
def index():
    errore = None

    if request.method == 'POST':
        query = (request.form.get('numero') or '').strip()
        query_digits = _digits(query)

        if query_digits.isdigit() and len(query_digits) == 5:
            risultato = cerca_voucher(query_digits)
            if risultato:
                return render_template('voucher.html', voucher=risultato, by_gift=False)
            else:
                errore = "Voucher non trovato. Controlla il numero inserito."

        elif query_digits.isdigit() and len(query_digits) == 4:
            try:
                sync_from_cloud()
                df = pd.read_excel(EXCEL_PATH)
            except Exception as e:
                return f"Errore lettura Excel: {e}"

            df.columns = df.columns.str.strip()

            if 'N° CARD' not in df.columns or 'ORDINE' not in df.columns:
                return "Colonne 'N° CARD' o 'ORDINE' non trovate nel file."

            q = query_digits.zfill(4)

            def _norm4(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return ""
                s = str(v).strip()
                if s.endswith('.0'):
                    s = s[:-2]
                if s.isdigit():
                    s = s.zfill(4)
                return s

            cards = df['N° CARD'].apply(_norm4)
            sel = cards == q

            if sel.any():
                ordine_val = str(df.loc[sel, 'ORDINE'].iloc[0]).strip()
                ordine_digits = _digits(ordine_val)
                risultato = cerca_voucher(ordine_digits)
                if risultato:
                    return render_template('voucher.html', voucher=risultato, by_gift=True)
                else:
                    errore = "Si è verificato un problema nel recupero della gift."
            else:
                errore = "Gift non ancora assegnata."

        else:
            errore = "Inserisci 5 cifre (voucher) oppure 4 cifre (gift)."

    return render_template('index.html', errore=errore)

@app.route('/gestisci', methods=['GET', 'POST'])
def gestisci():
    numero = ((request.args.get('numero') or request.form.get('numero')) or '').strip()
    if not numero:
        return "Numero voucher mancante"

    numero_digits = _digits(numero)

    try:
        sync_from_cloud()
        df = pd.read_excel(EXCEL_PATH)
    except Exception as e:
        return f"Errore lettura Excel: {e}"

    df.columns = df.columns.str.strip()
    sel = df['ORDINE'].astype(str).apply(_digits) == numero_digits
    if not sel.any():
        return "Voucher non trovato"

    index = df[sel].index[0]
    r = df.loc[index]

    serv_col = find_service_col(df.columns)

    # prima colonna scalatura libera
    prossimo = None
    for i, col in enumerate(['1', '2', '3', '4', '5'], start=1):
        val = r[col]
        if pd.isna(val) or (isinstance(val, str) and val.strip() == ""):
            prossimo = i
            break

    non_utilizzabile = (prossimo is None)
    label_appuntamento = f"Appuntamento {prossimo}" if prossimo else ""

    if request.method == 'POST':
        if non_utilizzabile:
            return "Voucher non più utilizzabile (scalature esaurite)"

        wb = None
        try:
            sync_from_cloud()
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            excel_row = index + 2  # +1 header

            # mappa header -> indice colonna
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            header_idx = {}
            for i, cell in enumerate(header_cells, start=1):
                key = (cell.value.strip() if isinstance(cell.value, str) else str(cell.value)) if cell.value is not None else ""
                header_idx[key] = i

            idx_scal = {str(i): header_idx.get(str(i)) for i in range(1, 6)}
            idx_card = header_idx.get('N° CARD')

            # aggiorna card (sempre, se il form ha inviato un valore)
            card_from_form = (request.form.get('card') or '').strip()
            if idx_card and card_from_form:
                ws[f'{get_column_letter(idx_card)}{excel_row}'] = card_from_form

            # colonna target della scalatura corrente
            col_idx = idx_scal.get(str(prossimo))
            if not col_idx:
                if wb: wb.close()
                return "Struttura file non valida: colonne 1-5 non trovate."
            col_letter = get_column_letter(col_idx)
            target_addr = f'{col_letter}{excel_row}'

            # lettura sicura del servizio (NaN/None/"" => vuoto)
            raw_serv = r.get(serv_col) if serv_col else None
            if raw_serv is None or (isinstance(raw_serv, float) and pd.isna(raw_serv)):
                has_service = False
            else:
                has_service = (str(raw_serv).strip() != "")

            # normalizza input importo
            importo_txt = (request.form.get('scalatura') or '').strip()
            importo_txt_norm = importo_txt.replace('€', '').replace('\xa0', ' ').strip().replace(',', '.')
            imp = _parse_money(importo_txt_norm)

            wrote = False
            if has_service:
                if request.form.get('servizio_effettuato'):
                    valore = _parse_money(r['VALORE']) or 0.0
                    ws[target_addr] = valore
                    wrote = True
            else:
                if imp is None or imp <= 0:
                    if wb: wb.close()
                    return "Importo non valido"
                ws[target_addr] = imp
                wrote = True

            if wrote:
                nota_form = (request.form.get('note') or '').strip()
                if nota_form:
                    cell = ws[target_addr]
                    prev = cell.comment.text if cell.comment else ""
                    prefix = "Servizio effettuato" if request.form.get('servizio_effettuato') else f"Appuntamento {prossimo}"
                    cell.comment = Comment(((prev + "\n") if prev else "") + f"{prefix}: {nota_form}", "WebApp")

            print("[DBG]",
                  "prossimo=", prossimo,
                  "target_addr=", target_addr,
                  "serv_pieno=", has_service,
                  "chk=", bool(request.form.get('servizio_effettuato')),
                  "imp_raw=", importo_txt,
                  "imp_norm=", importo_txt_norm,
                  "wrote=", wrote)

            wb.save(EXCEL_PATH)
            wb.close()

            ok = sync_to_cloud()
            return render_template('voucher.html', voucher=cerca_voucher(numero_digits, force_local=not ok), by_gift=False)

        except Exception as e:
            try:
                if wb:
                    wb.close()
            except:
                pass
            return f"Errore scrittura Excel: {e}"

    return render_template(
        'gestisci.html',
        label_appuntamento=label_appuntamento,
        numero=numero_digits,
        voucher={
            'numero': numero_digits,
            'card': (r['N° CARD'] if not pd.isna(r['N° CARD']) else ''),
            'note': '',
            'servizio': (r.get(serv_col) if (serv_col and not pd.isna(r.get(serv_col))) else ''),
            'valore': format_valore(r['VALORE']),
            'non_utilizzabile': non_utilizzabile,
            'prossimo_appuntamento': prossimo,
            'colonna_attiva': prossimo
        }
    )

# --- ROTTA ESISTENTE: assegna card a voucher già esistente ---
@app.route('/assegna-card', methods=['GET', 'POST'])
def assegna_card():
    numero = (request.args.get('numero') or '').strip()
    if not numero:
        return "Numero voucher mancante"

    numero_digits = _digits(numero)

    try:
        sync_from_cloud()
        df = pd.read_excel(EXCEL_PATH)
    except Exception as e:
        return f"Errore lettura Excel: {e}"

    df.columns = df.columns.str.strip()
    sel = df['ORDINE'].astype(str).apply(_digits) == numero_digits
    if not sel.any():
        return "Voucher non trovato"

    index = df[sel].index[0]
    r = df.loc[index]

    if request.method == 'POST':
        card_val = (request.form.get('card') or '').strip()
        if not card_val:
            return render_template(
                'assegna_card.html',
                numero=numero_digits,
                valore_card='',
                errore="Inserisci un numero card."
            )

        wb = None
        try:
            sync_from_cloud()
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            excel_row = index + 2
            ws[f'B{excel_row}'] = card_val
            wb.save(EXCEL_PATH)
            wb.close()

            ok = sync_to_cloud()
            return render_template('voucher.html', voucher=cerca_voucher(numero_digits, force_local=not ok), by_gift=False)

        except Exception as e:
            try:
                if wb:
                    wb.close()
            except:
                pass
            return f"Errore scrittura Excel: {e}"

    val_esistente = '' if pd.isna(r['N° CARD']) else str(r['N° CARD'])
    return render_template('assegna_card.html', numero=numero_digits, valore_card=val_esistente, errore=None)

# --- NUOVA ROTTA: apre la pagina assegna gift card ---
@app.route('/assegna-gift-card', methods=['GET', 'POST'])
def assegna_gift_card():
    form_data = {
        'gift_number': '',
        'importo': '',
        'cliente': '',
        'servizio': ''
    }
    errore = None

    if request.method == 'POST':
        form_data = {
            'gift_number': (request.form.get('gift_number') or '').strip(),
            'importo': (request.form.get('importo') or '').strip(),
            'cliente': (request.form.get('cliente') or '').strip(),
            'servizio': (request.form.get('servizio') or '').strip()
        }

        gift_digits = _digits(form_data['gift_number'])
        importo_val = _parse_money(form_data['importo'])
        cliente_val = form_data['cliente'].strip()
        servizio_val = form_data['servizio'].strip()

        if len(gift_digits) != 4:
            errore = "Il Numero Gift deve contenere esattamente 4 cifre."
            return render_template(
                'assegna_gift_card.html',
                errore=errore,
                form_data=form_data
            )

        if importo_val is None or importo_val <= 0:
            errore = "Inserisci un importo valido."
            return render_template(
                'assegna_gift_card.html',
                errore=errore,
                form_data=form_data
            )

        if not cliente_val:
            errore = "Il campo Cliente è obbligatorio."
            return render_template(
                'assegna_gift_card.html',
                errore=errore,
                form_data=form_data
            )

        wb = None
        try:
            sync_from_cloud()
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active

            if gift_number_exists(ws, gift_digits):
                wb.close()
                errore = "Questo numero Gift è già presente."
                return render_template(
                    'assegna_gift_card.html',
                    errore=errore,
                    form_data=form_data
                )

            # mappa header -> indice colonna
            header_cells = next(ws.iter_rows(min_row=1, max_row=1))
            header_idx = {}
            for i, cell in enumerate(header_cells, start=1):
                key = (cell.value.strip() if isinstance(cell.value, str) else str(cell.value)) if cell.value is not None else ""
                header_idx[key] = i

                        # colonne: usa header se trovati, altrimenti fallback fisso
            idx_ordine = header_idx.get('ORDINE') or 1
            idx_card = header_idx.get('N° CARD') or 2
            idx_cliente = header_idx.get('CLIENTE \\ MAIL ORDINE') or 3
            idx_data = header_idx.get('DATA') or 4
            idx_tipo = header_idx.get('TIPO') or 7
            idx_valore = header_idx.get('VALORE') or 8

            idx_servizio = None
            for k, v in header_idx.items():
                if k and str(k).strip().upper().startswith("SERVIZIO"):
                    idx_servizio = v
                    break
            idx_servizio = idx_servizio or 15

                        nuovo_ordine = get_next_manual_gift_order_number(ws)
            target_row = find_first_empty_row(ws, start_row=2, key_col=idx_ordine)

            # copia lo stile della riga precedente, se esiste
            if target_row > 2:
                copy_row_style(ws, target_row - 1, target_row, max_col=ws.max_column)

            ws.cell(row=target_row, column=idx_ordine).value = nuovo_ordine
            ws.cell(row=target_row, column=idx_card).value = gift_digits.zfill(4)
            ws.cell(row=target_row, column=idx_cliente).value = cliente_val
            ws.cell(row=target_row, column=idx_data).value = datetime.now().date()
            ws.cell(row=target_row, column=idx_tipo).value = "GIFT CARD"
            ws.cell(row=target_row, column=idx_valore).value = importo_val

            if servizio_val:
                ws.cell(row=target_row, column=idx_servizio).value = servizio_val

            wb.save(EXCEL_PATH)
            wb.close()

            ok = sync_to_cloud()
            nuovo_numero_digits = _digits(nuovo_ordine)

            return render_template(
                'voucher.html',
                voucher=cerca_voucher(nuovo_numero_digits, force_local=not ok),
                by_gift=False
            )

        except Exception as e:
            try:
                if wb:
                    wb.close()
            except:
                pass
            return f"Errore scrittura Excel: {e}"

    return render_template(
        'assegna_gift_card.html',
        errore=errore,
        form_data=form_data
    )

if __name__ == '__main__':
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "5000"))
    app.run(debug=DEV_MODE, host=host, port=port)